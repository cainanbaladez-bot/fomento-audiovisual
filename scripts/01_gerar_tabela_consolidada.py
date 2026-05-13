"""
gerar_tabela_consolidada.py
============================
Consolida dados de fomento público + desempenho por obra e exporta:
  - tabela_consolidada_obras.xlsx   (uma linha por obra)
  - tabela_consolidada_chamadas.xlsx (uma linha por chamada/categoria)

Fontes (todas locais, sem dependência da pasta codex):
  - obras_fomento_unificado.csv          → universo + CPB → NUMERO_SALIC / NUMERO_CONTRATO_FSA
  - projetos-fsa (1).csv                 → chamada + valores FSA (match por título)
  - projetos-com-renuncia-fiscal (2).csv → valores SALIC por artigo
  - crt_sinais_roi.csv                   → sinais de janelas (TV, VOD, DVD) por CPB
  - bilheteria .../agregado/por_filme_ano.csv → público por CPB×ano (2014+)
  - bilheteria_brasileira_consolidada.xlsx    → público pré-2014 por título
  - classificacao_chamadas_fsa.xlsx       → categoria (coluna C, revisada pelo usuário)
  - Festivais_por_obra.xlsx              → pontuação por festival (Oscar, Cannes, Berlim…)
  - lumiere_search.xlsx                  → admissões europeias (Lumière/CNC)
  - lumiere_vod_search.xlsx              → presença em VOD internacional

Uso:
    python gerar_tabela_consolidada.py
"""

import os, re, unicodedata, glob
from collections import defaultdict

import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Parâmetros ─────────────────────────────────────────────────────────────────
ANO_MAX = 2023   # exclui obras mais recentes (ciclo de vida incompleto)
PMI_2024 = 19.88  # preço médio ingresso 2024 (deflação constante)

PMI_SERIES = {      # preço médio ingresso por ano de exibição
    2009: 10.17, 2010: 10.67, 2011: 11.14, 2012: 11.46, 2013: 12.04,
    2014: 12.28, 2015: 13.59, 2016: 14.10, 2017: 15.00, 2018: 15.04,
    2019: 15.81, 2020: 15.94, 2021: 17.48, 2022: 19.10, 2023: 19.61,
    2024: 19.88, 2025: 19.88, 2026: 19.88,
}

TV_PAGA = {1: 280_000, 2: 120_000, 3: 60_000}   # R$ nominal (referência: ano de emissão do CRT)
VOD_DOM  = 60_000
TV_ABERTA = 80_000
DVD       = 5_000

# ── Deflator IPCA (base 2024) ──────────────────────────────────────────────────
IPCA_FATORES = {}
_ipca_path = os.path.join(ROOT, "tabelas_apoio", "deflator_ipca_base2024.csv")
try:
    _ipca_df = pd.read_csv(_ipca_path, sep=None, engine="python", encoding="utf-8-sig", dtype=str)
    for _, _r in _ipca_df.iterrows():
        try:
            _ano = int(float(str(_r.get("ano", "")).replace(",", ".")))
            _fat = float(str(_r.get("fator_real_2024", "1")).replace(",", "."))
            IPCA_FATORES[_ano] = _fat
        except Exception:
            pass
except Exception:
    pass  # sem deflator → fatores ficam vazios (ROI deflac = nominal)

# ── Labels canônicos de categoria
CAT_PONT_FEST_PROD   = "FSA Pontuação Festivais e Roteiro"
CAT_PONT_COM_PROD    = "FSA Pontuação Bilheteria e Roteiro — Produtora"
CAT_PONT_COM_DIST    = "FSA Pontuação Bilheteria e Roteiro — Distribuidora"
CAT_CONC_A           = "FSA Pontuação Bilheteria e Roteiro — Produtora"  # Módulo A → Comercial
CAT_CONC_B           = "FSA Pontuação Festivais e Roteiro"               # Módulo B → Festivais
CAT_AUTO             = "FSA Automático Bilheteria"
CAT_AUTO_DIST        = "FSA Automático Bilheteria"   # merged → mesmo label
CAT_AUTO_COM         = "FSA Automático Bilheteria"   # merged → mesmo label
CAT_AUTO_FEST        = "FSA Automático Festivais"
CAT_COPROD           = "FSA Coprodução Internacional"
CAT_SAV              = "FSA SAV/MINC / Arranjos Regionais"
CAT_ROTEIRO          = "FSA Apenas roteiro"
CAT_COMERCIALIZ      = "FSA Comercialização / Distribuição"
CAT_COMPL            = "FSA Complementação"
CAT_REN_ART3         = "Renúncia — Art.3/3-A/39"
CAT_REN_OUT          = "Renúncia — Outros Mecanismos"
CAT_TV               = "_tv_excluir"   # marcador interno — excluído do painel

# Limiar de valor FSA para desambiguar Módulo A (comercial) vs B (festivais) no CONCURSO
CONCURSO_MODULO_A_LIMIAR = 2_500_000

# Mapeamento explícito por nome de chamada → categoria
CATEGORIA_MAP_EXPLICT = {
    # ── Pontuação festivais ───────────────────────────────────────────────────
    "PRODECINE 03":   CAT_COMERCIALIZ,
    "PRODECINE 04":   CAT_COMPL,
    "PRODECINE 05":   CAT_PONT_FEST_PROD,
    # ── Pontuação comercial — produtora ──────────────────────────────────────
    "PRODECINE 01":   CAT_PONT_COM_PROD,
    "PRODECINE 06":   CAT_COPROD,
    "PRODUÇÃO CINEMA": CAT_PONT_COM_PROD,
    # ── Pontuação comercial — distribuidora ─────────────────────────────────
    "PRODECINE 02":                   CAT_PONT_COM_DIST,
    "CINEMA VIA DISTRIBUIDORA":       CAT_PONT_COM_DIST,
    "PRODUÇÃO CINEMA VIA DISTRIBUIDORA": CAT_PONT_COM_DIST,
    # ── Coprodução ────────────────────────────────────────────────────────────
    "PRODECINE 07":   CAT_COPROD,
    "PRODECINE 08":   CAT_COPROD,
    "PRODECINE 09":   CAT_COPROD,
    "PRODECINE 10":   CAT_COPROD,
    "PRODECINE 11":   CAT_COPROD,
    "PRODECINE 12":   CAT_COPROD,
    "COPRODUÇÃO INTERNACIONAL":         CAT_COPROD,
    "COPRODUÇÃO INTERNACIONAL CINEMA":  CAT_COPROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - COPRODUÇÃO CHILE-BRASIL":    CAT_COPROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - COPRODUÇÃO PORTUGAL-BRASIL": CAT_COPROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - COPRODUÇÃO URUGUAI-BRASIL":  CAT_COPROD,
    # ── Automático ────────────────────────────────────────────────────────────
    "FLUXO CONTÍNUO PRODUÇÃO PARA CINEMA":                CAT_AUTO,
    "FLUXO CONTÍNUO PRODUÇÃO PARA CINEMA - VIA DISTRIBUIDORA": CAT_AUTO_DIST,
    "SUPORTE AUTOMÁTICO - DESEMPENHO COMERCIAL CINEMA": CAT_AUTO_COM,
    "COMPLEMENTAÇÃO":                    CAT_COMPL,
    "COMERCIALIZAÇÃO EM CINEMA":         CAT_COMERCIALIZ,
    "COMERCIALIZAÇÃO - OPÇÃO DE INVESTIMENTO EM COMERCIALIZAÇÃO": CAT_COMERCIALIZ,
    "SUPORTE AUTOMÁTICO - DESEMPENHO ARTÍSTICO": CAT_AUTO_FEST,
    # ── SAV/MINC + Arranjos ───────────────────────────────────────────────────
    "SAV/MINC 01":  CAT_SAV, "SAV/MINC 02":  CAT_SAV, "SAV/MINC 03":  CAT_SAV,
    "SAV/MINC 04":  CAT_SAV, "SAV/MINC 05":  CAT_SAV, "SAV/MINC 06":  CAT_SAV,
    "SAV/MINC 07":  CAT_SAV, "SAV/MINC 08":  CAT_SAV, "SAV/MINC 09":  CAT_SAV,
    "SAV/MINC 10":  CAT_SAV, "SAV/MINC 11":  CAT_SAV, "SAV/MINC 13":  CAT_SAV,
    "ARRANJOS REGIONAIS": CAT_SAV,
    # ── Roteiro ───────────────────────────────────────────────────────────────
    "CINEMA NOVOS REALIZADORES": CAT_ROTEIRO,
    "CONCURSO PRODUÇÃO PARA CINEMA - MODULO A": CAT_PONT_COM_PROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - Modulo A": CAT_PONT_COM_PROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - MODULO B": CAT_PONT_FEST_PROD,
    "CONCURSO PRODUÇÃO PARA CINEMA - Modulo B": CAT_PONT_FEST_PROD,
    # ── TV / VOD — excluídos do painel de cinema ──────────────────────────────
    "SUPORTE AUTOMÁTICO - DESEMPENHO COMERCIAL TV E VOD": CAT_TV,
    "FLUXO CONTÍNUO PRODUÇÃO PARA TELEVISÃO":            CAT_TV,
    "PRODUÇÃO TV-VOD":                                   CAT_TV,
    "PRODUÇÃO TV/VOD - NOVOS REALIZADORES":              CAT_TV,
    "PRODUÇÃO TV/VOD – VIA PROGRAMADORA":                CAT_TV,
    "PRODAV 01":   CAT_TV, "PRODAV 02 - PROJETO DERIVADO": CAT_TV,
    "PRODAV 03":   CAT_TV, "PRODAV 05": CAT_TV, "PRODAV 06": CAT_TV,
    "PRODAV 07":   CAT_TV, "PRODAV 08": CAT_TV, "PRODAV 09": CAT_TV,
    "PRODAV 10":   CAT_TV, "PRODAV 11": CAT_TV, "PRODAV 12": CAT_TV,
    "PRODAV 13":   CAT_TV, "PRODAV 14": CAT_TV,
    "PRODAV - TVS PÚBLICAS": CAT_TV,
}

def get_categoria(chamada, valor_fsa=0):
    """Retorna categoria refinada. Para CONCURSO, usa valor_fsa como heurística de módulo."""
    if not chamada:
        return ""
    # Caso especial: CONCURSO — desambigua por valor
    if chamada.strip() == "CONCURSO PRODUÇÃO PARA CINEMA":
        return CAT_PONT_FEST_PROD
    cat = CATEGORIA_MAP_EXPLICT.get(chamada.strip())
    if cat:
        return cat
    upper = chamada.upper()
    # Padrões de fallback
    if any(p in upper for p in ["PRODAV", "TV-VOD", "TV/VOD", "TELEVISAO", "TELEVISÃO", "PRODUÇÃO TV"]):
        return CAT_TV
    if "COPROD" in upper:
        return CAT_COPROD
    if "SAV/MINC" in upper or "ARRANJOS" in upper:
        return CAT_SAV
    if "FLUXO CONTIN" in upper:
        return CAT_AUTO
    if "COMPLEMENT" in upper:
        return CAT_COMPL
    if "COMERCIALIZ" in upper:
        return CAT_COMERCIALIZ
    if "DESEMPENHO ART" in upper:
        return CAT_AUTO_FEST
    if "SUPORTE AUTO" in upper or "DESEMPENHO COM" in upper:
        return CAT_AUTO_COM
    return "outros_seletivos"

# ── Normalização de título ──────────────────────────────────────────────────────
def _norm(s):
    if not s:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# ── Parsing monetário ──────────────────────────────────────────────────────────
def _parse_money(s):
    if pd.isna(s):
        return 0.0
    s = str(s).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

# ── Universo de obras on-the-fly (substitui obras_fomento_unificado.csv) ───────
def _build_universo_fomento(root):
    """
    Full outer join de:
      obras-nao-pub-brasileiras-investimento-fsa (1).csv  → CPB + NUMERO_CONTRATO_FSA
      obras-nao-pub-brasileiras-fomento-indireto.csv      → CPB + NUMERO_SALIC
    Contratos/SALIC múltiplos por CPB são pipe-joined numa única linha.
    """
    fsa = pd.read_csv(
        os.path.join(root, "raw", "obras-nao-pub-brasileiras-investimento-fsa.csv"),
        sep=None, engine="python", encoding="utf-8-sig", dtype=str
    ).fillna("")
    fsa["CPB"] = fsa["CPB"].str.strip()
    fsa_agg = (
        fsa.groupby("CPB", sort=False)
        .agg(
            TITULO_ORIGINAL=("TITULO_ORIGINAL", "first"),
            NUMERO_CONTRATO_FSA=("NUMERO_CONTRATO_FSA",
                                  lambda x: "|".join(v for v in x if v)),
        ).reset_index()
    )
    fsa_agg["TEM_INVESTIMENTO_FSA"] = "SIM"

    ind = pd.read_csv(
        os.path.join(root, "raw", "obras-nao-pub-brasileiras-fomento-indireto.csv"),
        sep=None, engine="python", encoding="utf-8-sig", dtype=str
    ).fillna("")
    ind["CPB"] = ind["CPB"].str.strip()
    ind_agg = (
        ind.groupby("CPB", sort=False)
        .agg(
            TITULO_ORIGINAL=("TITULO_ORIGINAL", "first"),
            NUMERO_SALIC=("NUMERO_SALIC",
                           lambda x: "|".join(v for v in x if v)),
        ).reset_index()
    )
    ind_agg["TEM_FOMENTO_INDIRETO"] = "SIM"

    merged = fsa_agg.merge(ind_agg, on="CPB", how="outer", suffixes=("_fsa", "_ind"))
    merged["TITULO_ORIGINAL"] = merged["TITULO_ORIGINAL_fsa"].where(
        merged["TITULO_ORIGINAL_fsa"].notna() & (merged["TITULO_ORIGINAL_fsa"] != ""),
        merged["TITULO_ORIGINAL_ind"]
    )
    merged["NUMERO_CONTRATO_FSA"] = merged["NUMERO_CONTRATO_FSA"].fillna("")
    merged["TEM_INVESTIMENTO_FSA"] = merged["TEM_INVESTIMENTO_FSA"].fillna("NÃO")
    merged["NUMERO_SALIC"] = merged["NUMERO_SALIC"].fillna("")
    merged["TEM_FOMENTO_INDIRETO"] = merged["TEM_FOMENTO_INDIRETO"].fillna("NÃO")
    return merged[["CPB", "TITULO_ORIGINAL", "NUMERO_CONTRATO_FSA",
                   "TEM_INVESTIMENTO_FSA", "NUMERO_SALIC", "TEM_FOMENTO_INDIRETO"]]

# ═══════════════════════════════════════════════════════════════════════════════
# 1. UNIVERSO DE OBRAS
# ═══════════════════════════════════════════════════════════════════════════════
print("Construindo universo de obras a partir dos arquivos ANCINE brutos...")
unificado = _build_universo_fomento(ROOT)

# Normaliza colunas
unificado.columns = [c.strip() for c in unificado.columns]
unificado["CPB"] = unificado["CPB"].str.strip()
unificado["TITULO_NORM"] = unificado["TITULO_ORIGINAL"].apply(_norm)
unificado["TEM_FSA"]     = unificado["TEM_INVESTIMENTO_FSA"].str.strip().str.upper() == "SIM"
unificado["TEM_INDIR"]   = unificado["TEM_FOMENTO_INDIRETO"].str.strip().str.upper() == "SIM"

print(f"  Total de obras: {len(unificado)}")
print(f"  Com FSA: {unificado['TEM_FSA'].sum()}")
print(f"  Com renúncia/indireto: {unificado['TEM_INDIR'].sum()}")

# ── Ano de produção (dos arquivos anuais) ──────────────────────────────────────
print("Carregando anos de produção...")
obras_files = sorted(glob.glob(os.path.join(ROOT, "raw", "obras-nao-pub-brasileiras-csv", "*.csv")))
ano_frames = []
for f in obras_files:
    try:
        df = pd.read_csv(f, sep=None, engine="python", encoding="utf-8-sig", dtype=str,
                         usecols=lambda c: c in ["CPB", "ANO_PRODUCAO_INICIAL", "ANO_PRODUCAO_FINAL"])
        ano_frames.append(df)
    except Exception:
        pass
if ano_frames:
    ano_df = pd.concat(ano_frames, ignore_index=True).drop_duplicates("CPB")
    ano_df["ANO"] = pd.to_numeric(ano_df.get("ANO_PRODUCAO_INICIAL", pd.Series(dtype=str)),
                                   errors="coerce")
    cpb_to_ano = dict(zip(ano_df["CPB"], ano_df["ANO"]))
else:
    cpb_to_ano = {}

unificado["ANO"] = unificado["CPB"].map(cpb_to_ano)

# ═══════════════════════════════════════════════════════════════════════════════
# 2. FSA — valores e chamada por obra
# ═══════════════════════════════════════════════════════════════════════════════
# ── Aliases de título: título final da obra → título do projeto no FSA ─────────
# Usado quando o filme foi lançado com nome diferente do registrado no edital.
# Arquivo: tabelas_apoio/titulo_aliases_fsa.csv  (colunas: titulo_final, titulo_projeto_fsa)
_TITULO_ALIAS = {}   # titulo_norm_final → titulo_norm_projeto
_alias_path = os.path.join(ROOT, "tabelas_apoio", "titulo_aliases_fsa.csv")
try:
    _alias_df = pd.read_csv(_alias_path, encoding="utf-8-sig", dtype=str).fillna("")
    for _, _ar in _alias_df.iterrows():
        _tf = _norm(str(_ar.get("titulo_final", "")).strip())
        _tp = _norm(str(_ar.get("titulo_projeto_fsa", "")).strip())
        if _tf and _tp:
            _TITULO_ALIAS[_tf] = _tp
    print(f"  Aliases de título FSA carregados: {len(_TITULO_ALIAS)}")
except Exception:
    pass

print("Carregando projetos FSA...")
fsa_proj = pd.read_csv(
    os.path.join(ROOT, "raw", "projetos-fsa.csv"),
    sep=None, engine="python", encoding="utf-8-sig", dtype=str
).fillna("")
fsa_proj.columns = [c.strip() for c in fsa_proj.columns]
fsa_proj["TITULO_NORM"]  = fsa_proj["TITULO_PROJETO"].apply(_norm)
fsa_proj["VALOR_FSA"]    = fsa_proj.apply(
    lambda r: max(_parse_money(r.get("VALOR_TOTAL_LIBERADO", 0)),
                  _parse_money(r.get("VALOR_CONTRATO_DOU", 0))), axis=1)
fsa_proj["CHAMADA"]      = fsa_proj["CHAMADA_PUBLICA"].str.strip()

# ── Ano de deflação FSA (DATA_PRIMEIRO_DESEMBOLSO → fallback ANO_EDITAL) ──────
def _parse_ano_data(s):
    if not s or s == "" or (hasattr(s, "__class__") and s.__class__.__name__ == "float"):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return int(pd.to_datetime(str(s).strip(), format=fmt).year)
        except Exception:
            pass
    try:
        return int(pd.to_datetime(str(s).strip()).year)
    except Exception:
        return None

fsa_proj["_ANO_DEFLAC"] = fsa_proj.get("DATA_PRIMEIRO_DESEMBOLSO", pd.Series(dtype=str)).apply(_parse_ano_data)
_edital_num = pd.to_numeric(fsa_proj.get("ANO_EDITAL", pd.Series(dtype=str)), errors="coerce")
fsa_proj["_ANO_DEFLAC"] = fsa_proj["_ANO_DEFLAC"].where(
    fsa_proj["_ANO_DEFLAC"].notna(), _edital_num
)
fsa_ano_deflac = {}
for _, _r in fsa_proj.iterrows():
    _tn = _r.get("TITULO_NORM", "")
    _ano = _r.get("_ANO_DEFLAC")
    if _tn and _ano and not pd.isna(_ano) and _tn not in fsa_ano_deflac:
        fsa_ano_deflac[_tn] = int(_ano)

# Índice título → lista de projetos FSA (exact norm)
fsa_by_title = defaultdict(list)
# Índice de prefixo (primeiras 4 palavras) → lista de projetos FSA (fuzzy fallback)
fsa_by_prefix = defaultdict(list)

def _prefix4(s):
    return " ".join(s.split()[:4])

for _, row in fsa_proj.iterrows():
    tn = row["TITULO_NORM"]
    fsa_by_title[tn].append(row)
    fsa_by_prefix[_prefix4(tn)].append(row)

# Obras do FLUXO CONTÍNUO CINEMA onde o proponente é distribuidora (CNPJ diferente da produtora)
_fluxo_dist_norms = set()
for _, _r in fsa_proj[fsa_proj["CHAMADA"] == "FLUXO CONTÍNUO PRODUÇÃO PARA CINEMA"].iterrows():
    _cp = re.sub(r"\D", "", str(_r.get("CNPJ_PROPONENTE", "")))
    _cd = re.sub(r"\D", "", str(_r.get("CNPJ_PRODUTORA",  "")))
    if _cp and _cd and _cp != _cd:
        _fluxo_dist_norms.add(_r["TITULO_NORM"])
print(f"  FLUXO CONTÍNUO via distribuidora: {len(_fluxo_dist_norms)} títulos")

def _best_candidate(candidates, cnpj_requerente=""):
    if len(candidates) == 1:
        return candidates[0]["CHAMADA"], candidates[0]["VALOR_FSA"]
    if cnpj_requerente:
        cnpj_clean = re.sub(r"\D", "", cnpj_requerente)
        for r in candidates:
            if re.sub(r"\D", "", r.get("CNPJ_PRODUTORA", "")) == cnpj_clean:
                return r["CHAMADA"], r["VALOR_FSA"]
            if re.sub(r"\D", "", r.get("CNPJ_PROPONENTE", "")) == cnpj_clean:
                return r["CHAMADA"], r["VALOR_FSA"]
    best = max(candidates, key=lambda r: r["VALOR_FSA"])
    return best["CHAMADA"], best["VALOR_FSA"]

def match_fsa(titulo_norm, cnpj_requerente=""):
    """Retorna (chamada, valor_fsa) para a obra, ou ("", 0) se não casou."""
    # 1. Match exato por título normalizado
    candidates = fsa_by_title.get(titulo_norm, [])
    if candidates:
        return _best_candidate(candidates, cnpj_requerente)
    # 2. Match por prefixo de 4 palavras (recupera truncamentos e variações leves)
    prefix = _prefix4(titulo_norm)
    if len(prefix.split()) >= 3:  # exige pelo menos 3 palavras para evitar falsos positivos
        candidates = fsa_by_prefix.get(prefix, [])
        if candidates:
            return _best_candidate(candidates, cnpj_requerente)
    # 3. Match por containment: título da obra contido no título FSA ou vice-versa
    words_obra = set(titulo_norm.split())
    if len(words_obra) >= 3:
        for fsa_norm, fsa_list in fsa_by_title.items():
            words_fsa = set(fsa_norm.split())
            # Sobreposição de palavras ≥ 80% do menor título
            overlap = len(words_obra & words_fsa)
            min_len = min(len(words_obra), len(words_fsa))
            if min_len >= 3 and overlap / min_len >= 0.80:
                return _best_candidate(fsa_list, cnpj_requerente)
    return "", 0.0

def match_fsa_all(titulo_norm, cnpj_requerente=""):
    """
    Retorna lista de (chamada, valor) para TODAS as chamadas que a obra recebeu.
    Agrega por chamada (max valor por chamada) e ordena por valor desc.
    """
    # Aplica alias de título (ex.: filme lançado com nome diferente do projeto FSA)
    titulo_norm = _TITULO_ALIAS.get(titulo_norm, titulo_norm)
    # Coleta candidatos brutos por todos os métodos de match
    raw = []
    exact = fsa_by_title.get(titulo_norm, [])
    if exact:
        raw.extend(exact)
    else:
        prefix = _prefix4(titulo_norm)
        if len(prefix.split()) >= 3:
            raw.extend(fsa_by_prefix.get(prefix, []))
        if not raw:
            words_obra = set(titulo_norm.split())
            if len(words_obra) >= 3:
                for fsa_norm, fsa_list in fsa_by_title.items():
                    words_fsa = set(fsa_norm.split())
                    overlap = len(words_obra & words_fsa)
                    min_len = min(len(words_obra), len(words_fsa))
                    if min_len >= 3 and overlap / min_len >= 0.80:
                        raw.extend(fsa_list)
    if not raw:
        return []

    # Agrupa por CHAMADA → max VALOR_FSA
    by_chamada = defaultdict(float)
    for r in raw:
        by_chamada[r["CHAMADA"]] = max(by_chamada[r["CHAMADA"]], r["VALOR_FSA"])

    # Refinamento CNPJ: se alguma chamada bate o CNPJ, descarta as que não batem
    if cnpj_requerente:
        cnpj_clean = re.sub(r"\D", "", cnpj_requerente)
        chamadas_com_cnpj = set()
        for r in raw:
            r_cnpj = re.sub(r"\D", "", r.get("CNPJ_PRODUTORA", "") or r.get("CNPJ_PROPONENTE", ""))
            if r_cnpj == cnpj_clean:
                chamadas_com_cnpj.add(r["CHAMADA"])
        if chamadas_com_cnpj:
            by_chamada = {k: v for k, v in by_chamada.items() if k in chamadas_com_cnpj}

    # Desambigua FLUXO CONTÍNUO CINEMA: produtora vs distribuidora
    _FC  = "FLUXO CONTÍNUO PRODUÇÃO PARA CINEMA"
    _FCD = "FLUXO CONTÍNUO PRODUÇÃO PARA CINEMA - VIA DISTRIBUIDORA"
    if _FC in by_chamada and titulo_norm in _fluxo_dist_norms:
        by_chamada[_FCD] = by_chamada.pop(_FC)

    return sorted(by_chamada.items(), key=lambda x: x[1], reverse=True)

# Carrega CNPJ_REQUERENTE das obras para desambiguar
print("Carregando CNPJ das obras para desambiguação...")
cnpj_frames = []
for f in obras_files:
    try:
        df = pd.read_csv(f, sep=None, engine="python", encoding="utf-8-sig", dtype=str,
                         usecols=lambda c: c in ["CPB", "CNPJ_REQUERENTE"])
        cnpj_frames.append(df)
    except Exception:
        pass
cnpj_df = pd.concat(cnpj_frames, ignore_index=True).drop_duplicates("CPB") if cnpj_frames else pd.DataFrame(columns=["CPB","CNPJ_REQUERENTE"])
cpb_to_cnpj = dict(zip(cnpj_df["CPB"], cnpj_df["CNPJ_REQUERENTE"].fillna("")))

# ═══════════════════════════════════════════════════════════════════════════════
# 3. RENÚNCIA FISCAL — match por título + produtora (proxy, sem link CPB-SALIC)
#
# Arquivo esperado: projetos-com-renuncia-fiscal (2).csv  (ou variante)
# Colunas mínimas necessárias:
#   - título do projeto   → TITULO_PROJETO | NOME_PROJETO | TITULO (qualquer)
#   - CNPJ produtora      → CNPJ_PRODUTORA | CNPJ_PROPONENTE (opcional, refina match)
#   - valores por artigo  → CAPTADO_ART3, CAPTADO_ART3A [principal], + outros
#     (aceita prefixo "Soma de " e colunas com "ART3" no nome)
# ═══════════════════════════════════════════════════════════════════════════════
print("Carregando projetos com renuncia fiscal...")
_salic_path = os.path.join(ROOT, "raw", "projetos-com-renuncia-fiscal.csv")
_salic_loaded = False
salic = pd.DataFrame()
for _enc in ("utf-8-sig", "latin1", "cp1252"):
    try:
        salic = pd.read_csv(_salic_path, sep=None, engine="python",
                            encoding=_enc, dtype=str).fillna("")
        _salic_loaded = True
        break
    except UnicodeDecodeError:
        continue

art3_cols   = ["CAPTADO_ART3", "CAPTADO_ART3A", "CAPTADO_ART39"]
outros_cols = ["CAPTADO_ART1", "CAPTADO_ART1A", "CAPTADO_ART18", "CAPTADO_ART25",
               "CAPTADO_FUNCINES", "CAPTADO_EDITAL_ANCINE", "CAPTADO_PAR",
               "CAPTADO_PAQ", "CAPTADO_OUTROS_EDITAIS", "CAPTADO_LEI_ESTADUAL",
               "CAPTADO_LEI_MUNICIPAL", "CAPTADO_OUTRAS_FONTES",
               "CAPTADO_CONTRAPARTIDA", "CAPTADO_CONVERSAO"]

# Lookup: titulo_norm → {"art3", "outros", "cnpj"}  (para match por título+produtora)
ren_by_title   = defaultdict(list)   # titulo_norm → list of records
ren_by_prefix  = defaultdict(list)   # 4-word prefix → list of records

_ren_matched = 0

if _salic_loaded and not salic.empty:
    salic.columns = [c.strip() for c in salic.columns]

    # Normaliza "Soma de CAPTADO_ART3" → "CAPTADO_ART3"
    _col_map = {c: c.upper().replace("SOMA DE ", "").strip()
                for c in salic.columns if "SOMA DE " in c.upper()}
    salic.rename(columns=_col_map, inplace=True)

    # Detecta coluna de título (prioridade: TITULO_PROJETO > NOME_PROJETO > TITULO > 1ª col)
    _titulo_col = next(
        (c for c in salic.columns if any(k in c.upper()
         for k in ["TITULO_PROJETO", "NOME_PROJETO", "TITULO", "PROJETO"])),
        salic.columns[0]
    )
    # Detecta coluna de CNPJ produtora (opcional)
    _cnpj_col = next(
        (c for c in salic.columns if any(k in c.upper()
         for k in ["CNPJ_PRODUTORA", "CNPJ_PROPONENTE", "CNPJ"])),
        None
    )
    print(f"  Col titulo: '{_titulo_col}' | Col CNPJ: '{_cnpj_col}' | "
          f"Linhas: {len(salic)} | Cols: {list(salic.columns[:5])}")

    for _, row in salic.iterrows():
        _t = str(row.get(_titulo_col, "")).strip()
        if not _t or _t.isdigit() or _t.startswith("<"):
            continue   # pula linhas de totais/anos do pivot
        _tn = _norm(_t)
        if not _tn or len(_tn) < 3:
            continue
        _art3  = sum(_parse_money(row.get(c, 0)) for c in art3_cols   if c in salic.columns)
        _outros= sum(_parse_money(row.get(c, 0)) for c in outros_cols  if c in salic.columns)
        # Se nenhum valor por artigo, tenta coluna genérica de total
        if _art3 == 0 and _outros == 0:
            for _tc in ["TOTAL_CAPTADO", "VALOR_CAPTADO", "TOTAL"]:
                if _tc in salic.columns:
                    _art3 = _parse_money(row.get(_tc, 0))
                    break
        _rec = {"art3": _art3, "outros": _outros,
                "cnpj": re.sub(r"\D", "", str(row.get(_cnpj_col, ""))) if _cnpj_col else ""}
        ren_by_title[_tn].append(_rec)
        ren_by_prefix[_prefix4(_tn)].append(_rec)

    print(f"  Titulos renúncia carregados: {len(ren_by_title)}")
else:
    print("  [AVISO] projetos-com-renuncia-fiscal (2).csv nao encontrado ou vazio")
    print("          Renúncia = 0. Para ativar: arquivo precisa ter TITULO_PROJETO + CAPTADO_ART3/3A")


def _best_ren(candidates, cnpj_prod=""):
    """Escolhe melhor registro de renúncia dentre candidatos (prefere match CNPJ)."""
    if len(candidates) == 1:
        return candidates[0]
    if cnpj_prod:
        for r in candidates:
            if r["cnpj"] and r["cnpj"] == cnpj_prod:
                return r
    return max(candidates, key=lambda r: r["art3"] + r["outros"])


def get_salic_amounts(titulo_norm, cnpj_prod=""):
    """
    Retorna (art3, outros) via match por titulo_norm + CNPJ opcional.
    Estratégias em cascata: exato → prefixo 4 palavras → containment 80%.
    Prioridade: Art.3/3-A (lei do audiovisual) sobre outros mecanismos.
    """
    # 1. Match exato
    cands = ren_by_title.get(titulo_norm, [])
    if cands:
        r = _best_ren(cands, cnpj_prod)
        return r["art3"], r["outros"]
    # 2. Prefixo 4 palavras
    _pfx = _prefix4(titulo_norm)
    if len(_pfx.split()) >= 3:
        cands = ren_by_prefix.get(_pfx, [])
        if cands:
            r = _best_ren(cands, cnpj_prod)
            return r["art3"], r["outros"]
    # 3. Containment 80%
    _words = set(titulo_norm.split())
    if len(_words) >= 3:
        for _tn, _recs in ren_by_title.items():
            _wt = set(_tn.split())
            _ov = len(_words & _wt)
            _mn = min(len(_words), len(_wt))
            if _mn >= 3 and _ov / _mn >= 0.80:
                r = _best_ren(_recs, cnpj_prod)
                return r["art3"], r["outros"]
    return 0.0, 0.0

# ═══════════════════════════════════════════════════════════════════════════════
# 4. BILHETERIA
# ═══════════════════════════════════════════════════════════════════════════════
print("Carregando bilheteria...")

# 4a. 2014+ via dados diários agregados por CPB×ano
bil_ano_path = os.path.join(
    ROOT, "raw", "bilheteria-agregado", "por_filme_ano.csv")
bil_cpb = defaultdict(float)   # CPB → receita nominal total
bil_cpb_pub = defaultdict(float)  # CPB → público total

if os.path.exists(bil_ano_path):
    bil_df = pd.read_csv(bil_ano_path, sep=";", encoding="utf-8-sig", dtype=str)
    bil_df["PUBLICO"] = pd.to_numeric(bil_df["PUBLICO"], errors="coerce").fillna(0)
    bil_df["ANO"]     = pd.to_numeric(bil_df["ANO"], errors="coerce").fillna(0).astype(int)
    bil_df["CPB_ROE"] = bil_df["CPB_ROE"].str.strip()
    for _, row in bil_df.iterrows():
        cpb = row["CPB_ROE"]
        pub = row["PUBLICO"]
        ano = row["ANO"]
        pmi = PMI_SERIES.get(ano, PMI_2024)
        bil_cpb[cpb]     += pub * pmi
        bil_cpb_pub[cpb] += pub
    print(f"  CPBs com bilheteria (2014+): {len(bil_cpb)}")

# 4b. Pré-2014 via bilheteria_brasileira_consolidada.xlsx (match por título)
bil_pre = pd.read_excel(os.path.join(ROOT, "raw", "bilheteria_brasileira_consolidada.xlsx"))
bil_pre.columns = [c.strip() for c in bil_pre.columns]
bil_pre["ANO"] = pd.to_numeric(bil_pre.get("Ano Lançamento", pd.Series(dtype=str)),
                                errors="coerce").fillna(0).astype(int)
bil_pre["PUB"] = pd.to_numeric(bil_pre.get("Público Total", pd.Series(dtype=str)),
                                errors="coerce").fillna(0)
bil_pre["TITULO_NORM"] = bil_pre["Título"].apply(_norm)

# titulo_norm → (receita_nominal, público_total) para obras pré-2014
bil_pre_map = {}
for _, row in bil_pre.iterrows():
    ano = row["ANO"]
    pub = row["PUB"]
    pmi = PMI_SERIES.get(ano, 12.28)  # fallback conservador
    bil_pre_map[row["TITULO_NORM"]] = (pub * pmi, pub)

# ═══════════════════════════════════════════════════════════════════════════════
# 5. SINAIS CRT (janelas domésticas)
# ═══════════════════════════════════════════════════════════════════════════════
print("Carregando sinais CRT...")
crt = pd.read_csv(os.path.join(ROOT, "raw", "crt_sinais_roi.csv"),
                  sep=";", encoding="utf-8-sig", dtype=str).fillna("")
crt.columns = [c.strip() for c in crt.columns]
crt["CPB"] = crt["CPB"].str.strip()

def _safe_int_ano(v):
    try:
        return int(float(v)) if v and str(v).strip() not in ("", "nan") else None
    except Exception:
        return None

crt_map = {}
for _, row in crt.iterrows():
    cpb = row["CPB"]
    try:
        tier = int(float(row.get("tier_tv_paga", 0) or 0))
    except Exception:
        tier = 0
    crt_map[cpb] = {
        "tem_tv":              str(row.get("tem_tv_paga",   "0")).strip() == "1",
        "tier_tv":             tier,
        "tem_vod":             str(row.get("tem_vod",       "0")).strip() == "1",
        "tem_open":            str(row.get("tem_tv_aberta", "0")).strip() == "1",
        "tem_dvd":             str(row.get("tem_dvd",       "0")).strip() == "1",
        "ano_emissao_tv_paga": _safe_int_ano(row.get("ano_emissao_tv_paga")),
        "ano_emissao_vod":     _safe_int_ano(row.get("ano_emissao_vod")),
        "ano_emissao_tv_aberta": _safe_int_ano(row.get("ano_emissao_tv_aberta")),
        "ano_emissao_dvd":     _safe_int_ano(row.get("ano_emissao_dvd")),
    }

def estimar_janelas(cpb):
    """Retorna estimativa nominal de receita das janelas domésticas (ex-bilheteria)."""
    s = crt_map.get(cpb, {})
    rev = 0.0
    if s.get("tem_tv"):
        rev += TV_PAGA.get(s.get("tier_tv", 3), 60_000)
    if s.get("tem_vod"):
        rev += VOD_DOM
    if s.get("tem_open"):
        rev += TV_ABERTA
    if s.get("tem_dvd"):
        rev += DVD
    return rev


def estimar_janelas_deflac(cpb, ano_obra):
    """
    Retorna (outras_janelas_deflac_r2024, outras_janelas_nominal).
    Os valores base (TV_PAGA, VOD_DOM etc.) são nominais — referência do ano de emissão do CRT.
    Deflac = nominal * fator_ipca[ano_emissao] (converte para R$2024).
    """
    s = crt_map.get(cpb, {})
    deflac = 0.0
    nominal = 0.0

    def _janela(tem_key, valor_nominal, ano_emissao_key):
        if not s.get(tem_key):
            return 0.0, 0.0
        vn = float(valor_nominal)
        ano = s.get(ano_emissao_key) or ano_obra
        fat = IPCA_FATORES.get(int(ano), 1.0) if (ano and IPCA_FATORES) else 1.0
        return round(vn * fat, 2), vn

    tier = s.get("tier_tv", 3)
    vd, vn = _janela("tem_tv",  TV_PAGA.get(tier, 60_000), "ano_emissao_tv_paga")
    deflac += vd; nominal += vn
    vd, vn = _janela("tem_vod", VOD_DOM,   "ano_emissao_vod")
    deflac += vd; nominal += vn
    vd, vn = _janela("tem_open", TV_ABERTA, "ano_emissao_tv_aberta")
    deflac += vd; nominal += vn
    vd, vn = _janela("tem_dvd", DVD,        "ano_emissao_dvd")
    deflac += vd; nominal += vn
    return round(deflac, 2), round(nominal, 2)

# ═══════════════════════════════════════════════════════════════════════════════
# 6. DADOS INTERNACIONAIS
# ═══════════════════════════════════════════════════════════════════════════════
print("Carregando dados internacionais...")

import math as _math

# ── 6a. Festivais ──────────────────────────────────────────────────────────────
# Fonte única: resultados/festivais_consolidado.csv
#   Gerado por scripts/00b_fusao_festivais.py
#   Prioridade: Atas BRDE/FSA consolidadas (2014-2024, primária) +
#               Festivais_por_obra_pre_expansao.xlsx (complemento)

FEST_COLS = [
    "Oscar", "Cannes", "Berlim", "Veneza", "Sundance", "Locarno", "TIFF",
    "San Seb.", "Rotterdam", "Annecy", "Outros Intl", "Havana", "NYFF",
    "BFI London", "BAFTA", "Globo de Ouro", "Brasília", "Gramado",
    "Fest.Rio", "Mostra SP",
]

fest_map = {}     # titulo_norm (uppercase, script-01 _norm) → dict com pontos
fest_map_cpb = {} # cpb → dict (para filmes da Ata com CPB)

_consol_csv = os.path.join(ROOT, "resultados", "festivais_consolidado.csv")
try:
    _cdf = pd.read_csv(_consol_csv, encoding="utf-8-sig", dtype=str).fillna("")
    for _, _r in _cdf.iterrows():
        # Normaliza com a _norm() deste script (maiúsculas) para consistência.
        # 00b usa _norm() que produz minúsculas → sem re-normalização, get_fest() nunca acharia.
        _raw_titulo = _r.get("titulo", "") or _r.get("titulo_norm", "")
        _tn = _norm(str(_raw_titulo).strip())
        if not _tn:
            continue
        try:
            _tot = float(_r.get("pontuacao_total", "0").replace(",", "."))
        except Exception:
            _tot = 0.0
        _rec = {
            "Pontuação Festivais": _tot,
            "_fonte_festivais":    _r.get("fonte", ""),
            "_cpb_ata":            _r.get("cpb", ""),
        }
        for _fc in FEST_COLS:
            try:
                _v = float(str(_r.get(_fc, "0")).replace(",", "."))
            except Exception:
                _v = 0.0
            _rec[f"Festival — {_fc}"] = int(_v) if (_v and not _math.isnan(_v)) else 0
        fest_map[_tn] = _rec
        _cpb_ata = _r.get("cpb", "").strip()
        if _cpb_ata:
            fest_map_cpb[_cpb_ata] = _rec
    print(f"  Obras com dados de festivais: {len(fest_map)}")
except Exception as _e:
    print(f"  [AVISO] festivais_consolidado.csv: {_e}")

def get_fest(titulo_norm, cpb=""):
    """Retorna dados de festival pelo título (maiúsculo) ou CPB."""
    rec = fest_map.get(titulo_norm)
    if rec:
        return rec
    if cpb:
        return fest_map_cpb.get(cpb, {})
    return {}

# ── 6b. Lumière — admissões EU ─────────────────────────────────────────────────
lumiere_map = {}   # titulo_norm → admissões EU
try:
    lum = pd.read_excel(os.path.join(ROOT, "raw", "lumiere_search.xlsx"))
    lum.columns = [str(c).strip() for c in lum.columns]
    adm_col = next((c for c in lum.columns if "Adm" in c and "1996" in c), "Admissions 1996-2026")
    for _, row in lum.iterrows():
        t = row.get("Original title", "")
        if not isinstance(t, str):
            continue
        adm = pd.to_numeric(row.get(adm_col, 0), errors="coerce") or 0
        tn = _norm(t)
        lumiere_map[tn] = max(lumiere_map.get(tn, 0), int(adm))
    print(f"  Obras com Lumière: {len(lumiere_map)}")
except Exception as e:
    print(f"  [AVISO] Lumière: {e}")

def get_lumiere(titulo_norm):
    return lumiere_map.get(titulo_norm, 0)

# ── 6c. VOD Internacional ──────────────────────────────────────────────────────
vod_intl_map = {}   # titulo_norm → {"plataformas": N, "paises": N, "modelos": set}
try:
    vod_raw = pd.read_excel(os.path.join(ROOT, "raw", "lumiere_vod_search.xlsx"))
    vod_raw.columns = [str(c).strip() for c in vod_raw.columns]
    for _, row in vod_raw.iterrows():
        t = row.get("Original title", "")
        if not isinstance(t, str):
            continue
        tn = _norm(t)
        if tn not in vod_intl_map:
            vod_intl_map[tn] = {"plataformas": set(), "paises": set(), "modelos": set()}
        cat = str(row.get("Catalog", "")).strip()
        pais = str(row.get("Country", "")).strip()
        modelo = str(row.get("Business model", "")).strip()
        if cat:
            vod_intl_map[tn]["plataformas"].add(cat)
        if pais:
            vod_intl_map[tn]["paises"].add(pais)
        if modelo:
            vod_intl_map[tn]["modelos"].add(modelo)
    print(f"  Obras com VOD intl: {len(vod_intl_map)}")
except Exception as e:
    print(f"  [AVISO] VOD intl: {e}")

def get_vod_intl(titulo_norm):
    d = vod_intl_map.get(titulo_norm)
    if not d:
        return 0, 0, ""
    return len(d["plataformas"]), len(d["paises"]), "; ".join(sorted(d["modelos"]))

# ── 6e. Países alcançados ──────────────────────────────────────────────────────
paises_map = {}
try:
    _pp = pd.read_csv(
        os.path.join(ROOT, "tabelas_apoio", "paises_alcancados_apoio.csv"),
        sep=None, engine="python", encoding="utf-8-sig", dtype=str
    ).fillna("")
    for _, _r in _pp.iterrows():
        _cpb = str(_r.get("cpb", "")).strip()
        if _cpb:
            paises_map[_cpb] = {
                "total":      int(float(_r.get("paises_alcancados_total", 0) or 0)),
                "festivais":  _r.get("paises_festivais", ""),
                "lumiere":    _r.get("paises_lumiere_boxoffice_proxy", ""),
                "vod_europa": _r.get("paises_vod_europa", ""),
                "lista":      _r.get("paises_alcancados_lista", ""),
            }
    print(f"  Países por obra: {len(paises_map)}")
except Exception as _ep:
    print(f"  [AVISO] paises_alcancados_apoio.csv: {_ep}")

# ── 6d. ROI Internacional (composite 0–100) ────────────────────────────────────
_LUM_MAX = _math.log1p(2_500_000)   # normaliza log(1 + adm) pela maior obra conhecida
# Cap de pontuação de festivais para o componente de 70 pts do ROI Internacional.
# festivais_consolidado.csv reúne duas escalas distintas:
#   - Atas BRDE/FSA consolidadas 2014-2024 (primária): max ~115 pts (ANCINE)
#   - Festivais_por_obra_pre_expansao.xlsx (complemento): max ~351 pts
# Cap de 350 cobre o máximo observado (Central do Brasil, 351).
_FEST_MAX = 350

def roi_internacional(titulo_norm, cpb=""):
    """Score composto 0–100: 70% festivais + 20% Lumière + 10% VOD intl."""
    f_score = get_fest(titulo_norm, cpb).get("Pontuação Festivais", 0)
    l_adm   = get_lumiere(titulo_norm)
    _, v_pais, _ = get_vod_intl(titulo_norm)

    comp_fest = min(f_score / _FEST_MAX, 1.0) * 70
    comp_lum  = ((_math.log1p(l_adm) / _LUM_MAX) if l_adm > 0 else 0) * 20
    comp_vod  = (min(v_pais, 5) / 5) * 10   # satura em 5 países → 10 pts

    return round(comp_fest + comp_lum + comp_vod, 2)

print("Montando tabela por obra...")

rows = []
unmatched_fsa = 0

for _, obra in unificado.iterrows():
    cpb         = obra["CPB"]
    titulo      = obra["TITULO_ORIGINAL"].strip()
    titulo_norm = obra["TITULO_NORM"]
    ano         = obra["ANO"]
    tem_fsa     = obra["TEM_FSA"]
    tem_indir   = obra["TEM_INDIR"]
    cnpj        = cpb_to_cnpj.get(cpb, "")

    # Filtro de ano
    try:
        ano_int = int(ano)
    except (ValueError, TypeError):
        ano_int = 0
    if ano_int <= 0 or ano_int > ANO_MAX:
        continue

    # FSA — todas as chamadas da obra
    todas_chamadas_fsa = []   # lista de (chamada, valor)
    chamada   = ""
    valor_fsa = 0.0
    if tem_fsa:
        todas_chamadas_fsa = match_fsa_all(titulo_norm, cnpj)
        if todas_chamadas_fsa:
            chamada    = todas_chamadas_fsa[0][0]              # chamada primária = maior valor
            valor_fsa  = sum(v for _, v in todas_chamadas_fsa) # soma TODOS os contratos
        else:
            unmatched_fsa += 1

    todas_chamadas_str = "|".join(c for c, v in todas_chamadas_fsa)
    todos_valores_str  = "|".join(str(round(v, 2)) for c, v in todas_chamadas_fsa)

    # Renúncia fiscal (match por título + CNPJ)
    art3_sum, outros_sum = get_salic_amounts(titulo_norm, cnpj)
    tem_ren_art3   = art3_sum   > 0
    tem_ren_outros = outros_sum > 0

    # Bilheteria
    bil_nom   = bil_cpb.get(cpb, 0.0)
    bil_pub   = bil_cpb_pub.get(cpb, 0.0)
    if bil_nom == 0:
        pre = bil_pre_map.get(titulo_norm)
        if pre:
            bil_nom, bil_pub = pre
    bil_defl = bil_pub * PMI_2024

    # Outras janelas
    outras_janelas = estimar_janelas(cpb)
    outras_janelas_deflac, outras_janelas_nominal = estimar_janelas_deflac(cpb, ano_int)

    # Deflação de investimento (IPCA)
    _ano_deflac_fsa = fsa_ano_deflac.get(titulo_norm, ano_int)
    _fat_fsa = IPCA_FATORES.get(_ano_deflac_fsa, 1.0) if IPCA_FATORES else 1.0
    valor_fsa_deflac = round(valor_fsa * _fat_fsa, 2)
    _fat_ren = IPCA_FATORES.get(ano_int, 1.0) if IPCA_FATORES else 1.0
    ren_art3_deflac   = round(art3_sum  * _fat_ren, 2)
    ren_outros_deflac = round(outros_sum * _fat_ren, 2)
    inv_total_deflac  = valor_fsa_deflac + ren_art3_deflac + ren_outros_deflac

    # ROI deflacionado
    _receita_deflac  = bil_defl + outras_janelas_deflac
    roi_fsa_deflac   = round(_receita_deflac / valor_fsa_deflac,  4) if valor_fsa_deflac  > 0 else None
    roi_total_deflac = round(_receita_deflac / inv_total_deflac,  4) if inv_total_deflac  > 0 else None

    # Países
    _pd_data = paises_map.get(cpb, {})

    # Categoria primária (compatibilidade)
    if chamada:
        categoria = get_categoria(chamada, valor_fsa)
    elif art3_sum > 0:
        categoria = CAT_REN_ART3
    elif outros_sum > 0:
        categoria = CAT_REN_OUT
    elif tem_indir:
        categoria = CAT_REN_ART3   # fallback: indireto sem valor → trata como art3
    elif tem_fsa:
        # FSA presente mas sem match em projetos-fsa.csv (ex.: aprovado via Ata Desempenho
        # Artístico, que não consta no projetos-fsa.csv) → inferir categoria pelo festival
        _fonte_fest = get_fest(titulo_norm, cpb).get("_fonte_festivais", "")
        categoria = CAT_AUTO_FEST if "Desempenho Art" in _fonte_fest else "sem_categoria"
    else:
        categoria = "sem_categoria"

    # Internacional
    fest_data          = get_fest(titulo_norm, cpb)
    adm_eu             = get_lumiere(titulo_norm)
    vod_plat, vod_pais, vod_modelos = get_vod_intl(titulo_norm)
    roi_intl           = roi_internacional(titulo_norm, cpb)

    row = {
        "CPB":                            cpb,
        "Projeto":                        titulo,
        "Ano":                            ano_int,
        "Chamada":                        chamada,
        "Categoria":                      categoria,
        # ── Fomento ──────────────────────────────────────────────────────────
        "Valor FSA (R$)":                 round(valor_fsa, 2),
        "Renúncia Art.3/3-A/39 (R$)":     round(art3_sum, 2),
        "Renúncia Outros Mec. (R$)":      round(outros_sum, 2),
        # ── Desempenho doméstico ──────────────────────────────────────────────
        "Bilheteria Nominal (R$)":        round(bil_nom, 2),
        "Bilheteria Deflac. (R$)":        round(bil_defl, 2),
        "Estimativa Outras Janelas (R$)": round(outras_janelas, 2),
        # ── Internacional — composite ─────────────────────────────────────────
        "ROI Internacional (0-100)":      roi_intl,
        # ── Internacional — Festivais ────────────────────────────────────────
        "Pontuação Festivais":            fest_data.get("Pontuação Festivais", 0),
        "Fonte Festivais":                fest_data.get("_fonte_festivais", ""),
        **{f"Festival — {fc}": fest_data.get(f"Festival — {fc}", 0) for fc in FEST_COLS},
        # ── Internacional — Lumière ───────────────────────────────────────────
        "Adm. EU — Lumière":              adm_eu,
        # ── Internacional — VOD ───────────────────────────────────────────────
        "VOD Intl — N Plataformas":       vod_plat,
        "VOD Intl — N Países":            vod_pais,
        "VOD Intl — Modelos":             vod_modelos,
        # ── Multi-chamada FSA ─────────────────────────────────────────────────
        "Todas Chamadas FSA":             todas_chamadas_str,
        "Todos Valores FSA":              todos_valores_str,
        # ── Renúncia — flags separadas ────────────────────────────────────────
        "Tem Renuncia Art3":              tem_ren_art3,
        "Tem Renuncia Outros":            tem_ren_outros,
        # ── Investimento deflacionado (IPCA, base R$2024) ─────────────────────
        "Valor FSA Deflac. (R$2024)":          valor_fsa_deflac,
        "Renúncia Total Deflac. (R$2024)":     round(ren_art3_deflac + ren_outros_deflac, 2),
        "Investimento Total Deflac. (R$2024)": inv_total_deflac,
        # ── Outras janelas deflacionadas ──────────────────────────────────────
        "Outras Janelas Deflac. (R$2024)":     outras_janelas_deflac,
        "Outras Janelas Nominal (R$)":         outras_janelas_nominal,
        # ── ROI deflacionado ─────────────────────────────────────────────────
        "ROI Dom. FSA (deflac)":               roi_fsa_deflac,
        "ROI Dom. Total (deflac)":             roi_total_deflac,
        # ── Países alcançados ────────────────────────────────────────────────
        "Total Países Alcançados":             _pd_data.get("total", 0),
        "Países Festivais":                    _pd_data.get("festivais", ""),
        "Países Lumière":                      _pd_data.get("lumiere", ""),
        "Países VOD Europa":                   _pd_data.get("vod_europa", ""),
        "Países Lista":                        _pd_data.get("lista", ""),
    }
    rows.append(row)

df_obras = pd.DataFrame(rows)
print(f"  Obras na tabela: {len(df_obras)}")
print(f"  Com FSA sem match de chamada: {unmatched_fsa}")
print(f"  Com dados de festival: {(df_obras['Pontuação Festivais'] > 0).sum()}")
print(f"  Com Lumière: {(df_obras['Adm. EU — Lumière'] > 0).sum()}")
print(f"  Com VOD intl: {(df_obras['VOD Intl — N Plataformas'] > 0).sum()}")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. TABELA AGREGADA POR CHAMADA/CATEGORIA
# ═══════════════════════════════════════════════════════════════════════════════
print("Calculando agregados por chamada...")

def bayesian_agg(series, C, prior):
    """
    Média bayesiana: (n * media_bruta + C * prior) / (n + C).
    Retorna None se n == 0.
    """
    import pandas as _pd
    clean = _pd.to_numeric(series, errors="coerce").dropna()
    n = len(clean)
    if n == 0:
        return None
    return round((n * clean.mean() + C * prior) / (n + C), 4)


def agg_group(df_grp):
    import pandas as _pd_agg
    # ── Nominais (mantidos para compatibilidade) ──────────────────────────────
    inv_fsa   = df_grp["Valor FSA (R$)"].sum()
    ren_art3  = df_grp["Renúncia Art.3/3-A/39 (R$)"].sum()
    ren_out   = df_grp["Renúncia Outros Mec. (R$)"].sum()
    bil_nom   = df_grp["Bilheteria Nominal (R$)"].sum()
    bil_defl  = df_grp["Bilheteria Deflac. (R$)"].sum()
    outras_nom = df_grp["Estimativa Outras Janelas (R$)"].sum()

    # ── DEFLACIONADOS (R$2024) — usados em todos os ROIs ─────────────────────
    outras_deflac    = df_grp["Outras Janelas Deflac. (R$2024)"].sum()
    inv_fsa_deflac   = df_grp["Valor FSA Deflac. (R$2024)"].sum()
    inv_total_deflac = df_grp["Investimento Total Deflac. (R$2024)"].sum()
    receita_deflac   = bil_defl + outras_deflac

    # ROI agregado deflacionado (soma receita / soma investimento do grupo)
    roi_fsa_deflac   = round(receita_deflac / inv_fsa_deflac,   4) if inv_fsa_deflac   > 0 else None
    roi_total_deflac = round(receita_deflac / inv_total_deflac, 4) if inv_total_deflac > 0 else None

    # ── ROI por obra (deflac) — base para médias ──────────────────────────────
    roi_dom_col = _pd_agg.to_numeric(df_grp["ROI Dom. Total (deflac)"], errors="coerce")
    roi_fsa_col = _pd_agg.to_numeric(df_grp["ROI Dom. FSA (deflac)"],   errors="coerce")
    inv_w_col   = _pd_agg.to_numeric(df_grp["Investimento Total Deflac. (R$2024)"], errors="coerce").fillna(0)

    # Média simples (obras com ROI > 0)
    roi_dom_pos = roi_dom_col[roi_dom_col > 0]
    roi_dom_media_simples = round(roi_dom_pos.mean(), 4) if len(roi_dom_pos) > 0 else None

    # Média ponderada pelo investimento deflacionado (inclui zeros — ausência de retorno é resultado legítimo)
    mask_pos = (inv_w_col > 0)
    if mask_pos.sum() > 0:
        roi_dom_media_pond = round(
            (roi_dom_col[mask_pos] * inv_w_col[mask_pos]).sum() / inv_w_col[mask_pos].sum(), 4)
    else:
        roi_dom_media_pond = None

    # ROI Intl: média INCONDICIONAL (todas as obras, incluindo zeros)
    # Divisor = n total do grupo — comparável entre categorias com coberturas diferentes
    roi_intl_col = _pd_agg.to_numeric(df_grp["ROI Internacional (0-100)"], errors="coerce").fillna(0)
    n_com_intl_score = int((roi_intl_col >= 13).sum())
    roi_intl_media_simples = round(roi_intl_col.mean(), 2)
    roi_intl_media_pond = round(
        (roi_intl_col * inv_w_col).sum() / inv_w_col.sum(), 2) if inv_w_col.sum() > 0 else None

    # Países: média simples por obra
    paises_col = _pd_agg.to_numeric(df_grp["Total Países Alcançados"], errors="coerce").fillna(0)
    paises_media = round(paises_col.mean(), 2)

    # Métricas internacionais auxiliares
    n_com_festival = (df_grp["Pontuação Festivais"] > 0).sum()
    n_com_lumiere  = (df_grp["Adm. EU — Lumière"] > 0).sum()
    n_com_vod_intl = (df_grp["VOD Intl — N Plataformas"] > 0).sum()
    adm_eu_total   = df_grp["Adm. EU — Lumière"].sum()
    pont_fest_max  = df_grp["Pontuação Festivais"].max()
    pont_fest_med  = df_grp["Pontuação Festivais"].mean()

    return {
        "N Obras":                        len(df_grp),
        # ── Fomento ──────────────────────────────────────────────────────────
        "Valor FSA Total (R$)":                  round(inv_fsa, 0),
        "Valor FSA Deflac. Total (R$2024)":      round(inv_fsa_deflac, 0),
        "Renúncia Art.3/3-A/39 (R$)":            round(ren_art3, 0),
        "Renúncia Outros Mec. (R$)":             round(ren_out, 0),
        "Investimento Total Deflac. (R$2024)":   round(inv_total_deflac, 0),
        # ── Desempenho doméstico ──────────────────────────────────────────────
        "Bilheteria Nominal Total (R$)":         round(bil_nom, 0),
        "Bilheteria Deflac. Total (R$)":         round(bil_defl, 0),
        "Outras Janelas Deflac. Total (R$2024)": round(outras_deflac, 0),
        "Receita Estimada Deflac. (R$2024)":     round(receita_deflac, 0),
        # ── ROI doméstico deflacionado ────────────────────────────────────────
        "ROI Dom. Agregado (deflac, receita/inv_fsa)":   roi_fsa_deflac   if roi_fsa_deflac   is not None else "",
        "ROI Dom. Agregado (deflac, receita/inv_total)": roi_total_deflac if roi_total_deflac is not None else "",
        "ROI Dom. Média Simples (deflac)":               roi_dom_media_simples if roi_dom_media_simples is not None else "",
        "ROI Dom. Média Ponderada Inv. (deflac)":        roi_dom_media_pond    if roi_dom_media_pond    is not None else "",
        # ── Internacional — síntese ───────────────────────────────────────────
        "ROI Intl Média Simples (0-100)":        roi_intl_media_simples,
        "ROI Intl Média Ponderada Inv. (0-100)": roi_intl_media_pond if roi_intl_media_pond is not None else "",
        "ROI Intl Máximo (0-100)":               round(roi_intl_col.max(), 2),
        "N Obras c/ ROI Intl >= 13":             n_com_intl_score,
        "Países Alcançados Média":               paises_media,
        "N Obras c/ Festival":            int(n_com_festival),
        "N Obras c/ Lumière":             int(n_com_lumiere),
        "N Obras c/ VOD Intl":            int(n_com_vod_intl),
        "Adm. EU Total (Lumière)":        int(adm_eu_total),
        "Pontuação Festivais — Média":    round(pont_fest_med, 1),
        "Pontuação Festivais — Máximo":   int(pont_fest_max),
        "Índice Crítica (1-5)":           bayesian_agg(
            df_grp.get("CRITICA_INDICE_1_5", _pd_empty),
            C=10, prior=GLOBAL_PRIOR_CRITICA),
    }

# ── Crítica ──────────────────────────────────────────────────────────────────
import pandas as _pd_critica
_pd_empty = _pd_critica.Series(dtype=float)  # sentinela para colunas ausentes
critica_path = os.path.join(ROOT, "dados", "critica_obras.csv")
if os.path.exists(critica_path):
    df_critica = _pd_critica.read_csv(critica_path, dtype=str)
    df_critica["CRITICA_INDICE_1_5"] = _pd_critica.to_numeric(
        df_critica["CRITICA_INDICE_1_5"], errors="coerce")
    df_obras = df_obras.merge(
        df_critica[["CPB", "CRITICA_INDICE_1_5", "CRITICA_FONTES",
                    "CRITICA_N_FONTES", "CRITICA_CONFIANCA"]],
        on="CPB", how="left")
    print(f"  Critica_obras.csv integrado: "
          f"{df_obras['CRITICA_INDICE_1_5'].notna().sum()} obras com índice")
else:
    print("  AVISO: critica_obras.csv não encontrado — coluna CRITICA_INDICE_1_5 = None")
    for _col in ["CRITICA_INDICE_1_5", "CRITICA_FONTES",
                 "CRITICA_N_FONTES", "CRITICA_CONFIANCA"]:
        df_obras[_col] = None

GLOBAL_PRIOR_CRITICA = (
    _pd_critica.to_numeric(df_obras["CRITICA_INDICE_1_5"], errors="coerce").mean()
)
if _pd_critica.isna(GLOBAL_PRIOR_CRITICA):
    GLOBAL_PRIOR_CRITICA = 3.0  # prior neutro se nenhuma obra tem índice

# Por chamada (exclui: vazia → tratada como renúncia abaixo; _tv_excluir; sem_categoria)
chamada_rows = []
for chamada, grp in df_obras.groupby("Chamada"):
    if not chamada:
        continue  # obras sem chamada FSA → agregadas separadamente como renúncia
    cat = get_categoria(chamada)
    if cat in (CAT_TV, "sem_categoria"):
        continue
    row = {"Chamada": chamada, "Categoria": cat}
    row.update(agg_group(grp))
    chamada_rows.append(row)

# Renúncia (obras sem chamada FSA) — segmentada por mecanismo
ren_df = df_obras[df_obras["Chamada"] == ""]
if len(ren_df) > 0:
    ren_art3_df  = ren_df[ren_df["Renúncia Art.3/3-A/39 (R$)"] > 0]
    ren_out_df   = ren_df[ren_df["Renúncia Outros Mec. (R$)"] > 0]
    if len(ren_art3_df) > 0:
        row = {"Chamada": "Renúncia — Art.3/3-A/39", "Categoria": "renuncia_art3"}
        row.update(agg_group(ren_art3_df))
        chamada_rows.append(row)
    if len(ren_out_df) > 0:
        row = {"Chamada": "Renúncia — Outros Mecanismos", "Categoria": "renuncia_outros"}
        row.update(agg_group(ren_out_df))
        chamada_rows.append(row)

df_chamadas = pd.DataFrame(chamada_rows).sort_values("Receita Estimada Deflac. (R$2024)", ascending=False)

# ═══════════════════════════════════════════════════════════════════════════════
# 8. EXPORTAR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
print("Exportando Excel...")

_out_dir = os.path.join(ROOT, "resultados")
os.makedirs(_out_dir, exist_ok=True)
out_obras    = os.path.join(_out_dir, "tabela_consolidada_obras.xlsx")
out_chamadas = os.path.join(_out_dir, "tabela_consolidada_chamadas.xlsx")

with pd.ExcelWriter(out_obras, engine="openpyxl") as writer:
    # Aba obras — ordena por chamada > ano > título
    df_obras_exp = df_obras.sort_values(
        ["Chamada", "Ano", "Projeto"]).reset_index(drop=True)  # mantém CPB para cruzamento com diversidade
    df_obras_exp.to_excel(writer, sheet_name="Obras", index=False)

    # Aba chamadas
    df_chamadas.to_excel(writer, sheet_name="Chamadas", index=False)

    # Formata colunas monetárias
    wb = writer.book
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for sheetname in ["Obras", "Chamadas"]:
        ws = wb[sheetname]
        for col_cells in ws.iter_cols(min_row=1, max_row=1):
            cell = col_cells[0]
            cell.font = Font(bold=True)
        # Autofit aproximado
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

print(f"  Salvo: {out_obras}")

# Chamadas separadas
with pd.ExcelWriter(out_chamadas, engine="openpyxl") as writer:
    df_chamadas.to_excel(writer, sheet_name="Chamadas", index=False)
    wb = writer.book
    ws = wb["Chamadas"]
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        col_cells[0].font = Font(bold=True)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

print(f"  Salvo: {out_chamadas}")

# ── Resumo no terminal ──────────────────────────────────────────────────────────
print()
print("=" * 60)
print(f"Obras exportadas:    {len(df_obras)}")
print(f"Com bilheteria:      {(df_obras['Bilheteria Nominal (R$)'] > 0).sum()}")
print(f"Com FSA:             {(df_obras['Valor FSA (R$)'] > 0).sum()}")
print(f"Com renúncia art3:   {(df_obras['Renúncia Art.3/3-A/39 (R$)'] > 0).sum()}")
print(f"Com renúncia outros: {(df_obras['Renúncia Outros Mec. (R$)'] > 0).sum()}")
print()
print("Resumo por categoria (ROI = receita / FSA):")
for _, row in df_chamadas.head(20).iterrows():
    print(f"  {row['Categoria']:40s}  n={row['N Obras']:4d}  ROI={row['ROI Dom. Agregado (deflac, receita/inv_fsa)']}")
print("=" * 60)
print("Concluído.")
